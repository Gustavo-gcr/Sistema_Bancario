import random
import openpyxl

class Banco:
    def __init__(self, arquivo_excel):
        self.arquivo_excel = arquivo_excel
        try:
            self.wb = openpyxl.load_workbook(arquivo_excel)
            self.sheet = self.wb.active
            if self.sheet.max_row == 1:  # Verifica se o arquivo estava vazio e adiciona cabeçalho
                self.sheet.append(['IDADE', 'NOME', 'SALDO', 'NUM_CONTA'])
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
            self.sheet.append(['IDADE', 'NOME', 'SALDO', 'NUM_CONTA'])
            self.wb.save(self.arquivo_excel)

    def cadastrar_usuario(self, nome, idade, saldo):
        if idade < 18:
            print(f'Ola, infelizmente sua idade impede de você criar uma conta, já que você tem {idade} anos e a idade mínima exigida é de 18 anos.')
            return False

        numero_conta = random.randint(1, 999)
        nova_conta = {
            'IDADE': idade,
            'NOME': nome,
            'SALDO': saldo,
            'NUM_CONTA': numero_conta
        }
        self.sheet.append([idade, nome, saldo, numero_conta])
        self.wb.save(self.arquivo_excel)
        print(f'Conta cadastrada com sucesso! Número da conta: {numero_conta}')
        return nova_conta

    def depositar(self, numero_conta, valor):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            conta['SALDO'] += valor
            self.atualizar_conta(conta)
            print(f'Depósito de R${valor} realizado com sucesso! Saldo atual: R${conta["SALDO"]}')
        else:
            print('Conta não encontrada.')

    def pagar(self, numero_conta, valor):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            if conta['SALDO'] >= valor:
                conta['SALDO'] -= valor
                self.atualizar_conta(conta)
                print(f'Pagamento de R${valor} realizado com sucesso! Saldo atual: R${conta["SALDO"]}')
            else:
                print(f'Saldo insuficiente para realizar o pagamento. Saldo atual: R${conta["SALDO"]}')
        else:
            print('Conta não encontrada.')

    def ver_saldo(self, numero_conta):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            print(f'Saldo da conta {numero_conta}: R${conta["SALDO"]}')
        else:
            print('Conta não encontrada.')

    def ver_todas_contas(self):
        print("\n=== Lista de Contas Cadastradas ===")
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            print(f"Nome: {row[1]}, Número da conta: {row[3]}, Saldo: {'****' if row[2] else '****'}")

    def atualizar_conta(self, numero_conta, novo_nome=None, nova_idade=None, novo_saldo=None):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            if novo_nome:
                conta['NOME'] = novo_nome
            if nova_idade:
                conta['IDADE'] = nova_idade
            if novo_saldo:
                conta['SALDO'] = novo_saldo
            self.atualizar_conta_excel(conta)
            print(f'Conta {numero_conta} atualizada com sucesso!')
        else:
            print('Conta não encontrada.')

    def excluir_conta(self, numero_conta):
        conta = self.encontrar_conta(numero_conta)
        if conta:
            for row in self.sheet.iter_rows(min_row=2):
                if row[3].value == numero_conta:
                    self.sheet.delete_rows(row[0].row)
                    self.wb.save(self.arquivo_excel)
                    print(f'Conta {numero_conta} excluída com sucesso!')
                    return
            print('Erro ao excluir conta.')
        else:
            print('Conta não encontrada.')

    def encontrar_conta(self, numero_conta):
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[3] == numero_conta:
                return {
                    'IDADE': row[0],
                    'NOME': row[1],
                    'SALDO': row[2],
                    'NUM_CONTA': row[3]
                }
        return None

    def atualizar_conta_excel(self, conta):
        for row in self.sheet.iter_rows(min_row=2):
            if row[3].value == conta['NUM_CONTA']:
                row[0].value = conta['IDADE']
                row[1].value = conta['NOME']
                row[2].value = conta['SALDO']
                self.wb.save(self.arquivo_excel)
                break

# Função para exibir o menu
def exibir_menu():
    print("\n=== MENU ===")
    print("1. Cadastrar uma nova conta")
    print("2. Depositar Dinheiro")
    print("3. Realizar Pagamento")
    print("4. Ver saldo da conta")
    print("5. Ver todas as contas cadastradas")
    print("6. Atualizar uma conta")
    print("7. Excluir uma conta")
    print("0. Sair do programa")

def main():
    banco = Banco('contas.xlsx')

    while True:
        exibir_menu()
        opcao = input('Escolha uma opção: ')

        if opcao == '1':
            nome = input('Nome do titular: ')
            idade = int(input('Idade: '))
            saldo = float(input('Saldo inicial: R$'))
            banco.cadastrar_usuario(nome, idade, saldo)

        elif opcao == '2':
            numero_conta = int(input('Número da conta: '))
            valor = float(input('Valor para depósito: R$'))
            banco.depositar(numero_conta, valor)

        elif opcao == '3':
            numero_conta = int(input('Número da conta: '))
            valor = float(input('Valor para pagamento: R$'))
            banco.pagar(numero_conta, valor)

        elif opcao == '4':
            numero_conta = int(input('Número da conta: '))
            banco.ver_saldo(numero_conta)

        elif opcao == '5':
            banco.ver_todas_contas()

        elif opcao == '6':
            numero_conta = int(input('Número da conta: '))
            print("Deixe o campo em branco se não quiser atualizar.")
            novo_nome = input('Novo nome (opcional): ') or None
            nova_idade = input('Nova idade (opcional): ')
            nova_idade = int(nova_idade) if nova_idade else None
            novo_saldo = input('Novo saldo (opcional): ')
            novo_saldo = float(novo_saldo) if novo_saldo else None
            banco.atualizar_conta(numero_conta, novo_nome, nova_idade, novo_saldo)

        elif opcao == '7':
            numero_conta = int(input('Número da conta a ser excluída: '))
            banco.excluir_conta(numero_conta)

        elif opcao == '0':
            print('Programa encerrado. Obrigado por utilizar!')
            break
        else:
            print('Opção inválida. Escolha novamente.')

if __name__ == "__main__":
    main()
